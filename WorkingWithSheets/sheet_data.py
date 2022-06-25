import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}


def calc_product_per_supplier():
    # calculation number of products per supplier
    try:
        products_per_supplier[supplier_name] += 1
    except KeyError:
        products_per_supplier[supplier_name] = 1


def calc_total_value_per_supplier():
    # calculation total value of inventory per supplier
    try:
        total_value_per_supplier[supplier_name] += inventory * price
    except KeyError:
        total_value_per_supplier[supplier_name] = inventory * price


def find_inventory_under_10():
    # products with inventory under 10
    if inventory < 10:
        products_under_10[int(product_num)] = int(inventory)


def update_total_value():
    global inventory_price
    # add value for total inventory price
    inventory_price.value = inventory * price


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    calc_product_per_supplier()

    calc_total_value_per_supplier()

    find_inventory_under_10()

    update_total_value()

inv_file.save("inventory_with_total_value.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10)